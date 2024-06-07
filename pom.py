import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from Excel_utilities import getRowCount, readData, fillGreenColor, fillRedColor





class LoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(self.driver, 15)

    def login(self):
        file_path = "C:/Users/rsund/Desktop/test_data.xlsx"
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active  # Get the active sheet

            rows = getRowCount(file_path, sheet.title)
            for r in range(2, rows + 1):
                username = readData(file_path, sheet.title, r, 2)
                password = readData(file_path, sheet.title, r, 3)

                try:
                    self.wait.until(EC.presence_of_element_located((By.NAME, "username"))).clear()
                    self.wait.until(EC.visibility_of_element_located((By.NAME, "username"))).send_keys(username)

                    self.wait.until(EC.presence_of_element_located((By.NAME, "password"))).clear()
                    self.wait.until(EC.visibility_of_element_located((By.NAME, "password"))).send_keys(password)

                    login_button = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".oxd-button.oxd-button--medium.oxd-button--main.orangehrm-login-button")))
                    login_button.click()

                    # Check if the current URL contains "dashboard"
                    if "dashboard" in self.driver.current_url.lower():
                        print("Test Passed")
                        test_result = "Pass"
                        fillGreenColor(file_path, sheet.title, r, 7)
                    else:
                        print("Test Failed")
                        test_result = "Fail"
                        fillRedColor(file_path, sheet.title, r, 7)

                    sheet.cell(row=r, column=7, value=test_result)
                    sheet.cell(row=r, column=4, value=time.strftime("%Y-%m-%d"))
                    sheet.cell(row=r, column=5, value=time.strftime("%H:%M:%S"))
                    sheet.cell(row=r, column=6, value="Raja")

                except Exception as e:
                    print(f"Error during login for row {r}: {e}")
                    fillRedColor(file_path, sheet.title, r, 7)
                    sheet.cell(row=r, column=7, value="Error")

            workbook.save(file_path)

        except Exception as e:
            print(f"Error occurred: {e}")

        finally:
            workbook.close()
